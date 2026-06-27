# ── FITTRACK BOT — FINAL CLEAN VERSION ──
# Chenna Kesava Reddy
#
# HOW IT WORKS:
# ─────────────────────────────────────────────────────
# IDENTIFICATION:
#   1. Message from OWN_NUMBER         → ignore (bot's own message)
#   2. Message starts with "trainer"   → check TRAINER_NUMBERS → process trainer command
#   3. Sender in phone_mapping.json    → registered client → process client command
#   4. Sender not in mapping           → unregistered → tell them to contact trainer
#
# TRAINER COMMANDS (only from TRAINER_NUMBERS):
#   trainer register Rahul, 28, 82kg, 175cm, B+, Veg, phone 919533526497
#   trainer update RAH-001 weight 79kg
#   trainer note RAH-001 great session today
#   trainer progress RAH-001
#   trainer details RAH-001
#   trainer list
#
# CLIENT COMMANDS (only registered clients):
#   update weight 79kg
#   update goal muscle gain
#   my details
#   my progress
# ─────────────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()
from flask import Flask, request
import requests as req
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os, json, base64
from datetime import datetime

# ══════════════════════════════════════════════════════
# CONFIG — update these values
# ══════════════════════════════════════════════════════
ULTRAMSG_INSTANCE = "instance182370"
ULTRAMSG_TOKEN    = os.getenv("ULTRAMSG_TOKEN", "YOUR_ULTRAMSG_TOKEN_HERE")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "YOUR_ANTHROPIC_API_KEY_HERE")
OWN_NUMBER        = os.getenv("OWN_NUMBER", "917022206001")
DESKTOP_PATH      = r"C:\Users\mckre\OneDrive\Desktop\FitTrack_Clients"



# Your UltraMsg number — bot ignores messages from this number
OWN_NUMBER = "917022206001"
 
# Trainer phone numbers — only these can send trainer commands
TRAINER_NUMBERS = [
    "918123009006",   # your mobile
    "917022206001",   # UltraMsg number
]
 
# Derived URLs — no need to change
ULTRAMSG_CHAT_URL = "https://api.ultramsg.com/" + ULTRAMSG_INSTANCE + "/messages/chat"
ULTRAMSG_FILE_URL = "https://api.ultramsg.com/" + ULTRAMSG_INSTANCE + "/messages/document"
MAPPING_FILE      = os.path.join(DESKTOP_PATH, "phone_mapping.json")
 
# Excel column headers
HEADERS = [
    "Type", "Client ID", "Client Name", "Age", "Weight",
    "Height", "Blood Group", "Diet Type", "Fitness Goal",
    "Notes", "Date", "Updated By"
]
 
# ══════════════════════════════════════════════════════
# APP INIT
# ══════════════════════════════════════════════════════
app = Flask(__name__)
 
if not os.path.exists(DESKTOP_PATH):
    os.makedirs(DESKTOP_PATH)
    print("Created folder: " + DESKTOP_PATH)
 
# ══════════════════════════════════════════════════════
# MAPPING FUNCTIONS
# phone_mapping.json stores:
#   phones  : { "919533526497": "RAH-001" }
#   ids     : { "RAH-001": { name, filename, phone, registered } }
#   clients : { "rahul sharma": "RAH-001" }
#   prefixes: { "RAH": 1, "PRI": 2 }
# ══════════════════════════════════════════════════════
 
def load_mapping():
    if os.path.exists(MAPPING_FILE):
        with open(MAPPING_FILE, 'r') as f:
            return json.load(f)
    return {"phones": {}, "ids": {}, "clients": {}, "prefixes": {}}
 
def save_mapping(data):
    with open(MAPPING_FILE, 'w') as f:
        json.dump(data, f, indent=2)
 
def clean_number(phone):
    """Strip WhatsApp suffixes and normalize to include country code.
    Ensures 919916696004 and 9916696004 both resolve to 919916696004
    so trainer-registered numbers always match client incoming numbers."""
    # Strip WhatsApp suffixes
    number = phone.replace('@c.us','').replace('@lid','').split('@')[0].strip()
    # Remove any leading + sign
    number = number.lstrip('+')
    # Normalize Indian numbers — if 10 digits starting with 6-9, add 91
    if len(number) == 10 and number[0] in '6789':
        number = '91' + number
    return number
 
def generate_id(name):
    """RAH-001 for Rahul, PRI-001 for Priya, RAH-002 for second Rahul"""
    m      = load_mapping()
    prefix = name.strip().split()[0][:3].upper()
    count  = m.get("prefixes", {}).get(prefix, 0) + 1
    m.setdefault("prefixes", {})[prefix] = count
    save_mapping(m)
    return prefix + "-" + str(count).zfill(3)
 
def register_client(phone, client_id, name, filename):
    """Save phone → client_id mapping after registration"""
    m  = load_mapping()
    cp = clean_number(phone)
    m["phones"][cp]       = client_id
    m["clients"][name.lower()] = client_id
    m["ids"][client_id]   = {
        "name": name, "filename": filename,
        "phone": cp,
        "registered": datetime.now().strftime("%d-%b-%Y %H:%M")
    }
    save_mapping(m)
    print("Registered: " + cp + " → " + client_id)
 
def get_client_by_phone(sender):
    """Find client record using sender's phone number"""
    m   = load_mapping()
    cp  = clean_number(sender)
    cid = m.get("phones", {}).get(cp)
    if not cid:
        return None, None, None, None
    info = m.get("ids", {}).get(cid, {})
    fp   = os.path.join(DESKTOP_PATH, info.get("filename", ""))
    if os.path.exists(fp):
        return fp, info["filename"], cid, info["name"]
    return None, None, None, None
 
def get_client_by_id(client_id):
    """Find client record using client ID like RAH-001"""
    m    = load_mapping()
    cid  = client_id.upper().strip()
    info = m.get("ids", {}).get(cid, {})
    if not info:
        return None, None, None, None
    fp = os.path.join(DESKTOP_PATH, info.get("filename", ""))
    if os.path.exists(fp):
        return fp, info["filename"], cid, info["name"]
    return None, None, None, None
 
def get_client_by_name(name):
    """Find client record using client name (partial match supported)"""
    m  = load_mapping()
    nl = name.lower().strip()
    # exact match
    cid = m.get("clients", {}).get(nl)
    if cid:
        return get_client_by_id(cid)
    # partial match
    for stored_name, cid in m.get("clients", {}).items():
        if nl in stored_name or stored_name in nl:
            return get_client_by_id(cid)
    return None, None, None, None
 
def find_client(identifier):
    """Smart lookup — works with ID (RAH-001) or name (Rahul)"""
    identifier = identifier.strip()
    parts = identifier.upper().split('-')
    if len(parts) == 2 and parts[0].isalpha() and parts[1].isdigit():
        return get_client_by_id(identifier)
    return get_client_by_name(identifier)
 
def get_phone_for_client(client_id):
    """Get client's phone number from mapping"""
    m = load_mapping()
    return m.get("ids", {}).get(client_id.upper(), {}).get("phone", "")
 
def trainer_phone():
    """Return first non-UltraMsg trainer number"""
    for t in TRAINER_NUMBERS:
        if t != OWN_NUMBER:
            return t
    return TRAINER_NUMBERS[0]
 
def is_trainer(sender):
    """Check if sender is a registered trainer"""
    cp = clean_number(sender)
    return any(t in cp or cp in t for t in TRAINER_NUMBERS)
 
def is_registered(sender):
    """Check if sender is a registered client"""
    fp, *_ = get_client_by_phone(sender)
    return fp is not None
 
# ══════════════════════════════════════════════════════
# WHATSAPP FUNCTIONS
# ══════════════════════════════════════════════════════
 
def send_text(to, message):
    """Send a WhatsApp text message"""
    cp = clean_number(to)
    try:
        r = req.post(ULTRAMSG_CHAT_URL, timeout=10,
                     data={"token": ULTRAMSG_TOKEN, "to": cp, "body": message})
        print("Text → " + cp + " [" + str(r.status_code) + "]")
    except Exception as e:
        print("Text error: " + str(e))
 
def send_file(to, filepath, filename, caption):
    """Send a WhatsApp file attachment"""
    cp = clean_number(to)
    try:
        with open(filepath, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('utf-8')
        mime = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        r = req.post(ULTRAMSG_FILE_URL, timeout=15, data={
            "token": ULTRAMSG_TOKEN, "to": cp,
            "filename": filename, "caption": caption,
            "document": mime + b64
        })
        print("File → " + cp + " [" + str(r.status_code) + "]")
    except Exception as e:
        print("File error: " + str(e))
 
def send_to_trainer(message, fp=None, fn=None, caption=None):
    """Send message (and optionally file) to trainer"""
    send_text(trainer_phone(), message)
    if fp and fn and caption:
        send_file(trainer_phone(), fp, fn, caption)
 
def send_to_client(client_id, message, fp=None, fn=None, caption=None):
    """Send message (and optionally file) to client using their ID"""
    cp = get_phone_for_client(client_id)
    if not cp:
        print("No phone found for: " + client_id)
        return
    send_text(cp, message)
    if fp and fn and caption:
        send_file(cp, fp, fn, caption)
 
# ══════════════════════════════════════════════════════
# CLAUDE AI FUNCTIONS
# ══════════════════════════════════════════════════════
 
def ask_claude(prompt, max_tokens=300):
    """Send a prompt to Claude and return clean text"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    r = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}]
    )
    return r.content[0].text.strip().replace('```json','').replace('```','').strip()
 
def detect_client_intent(message):
    """Classify client message as update / my_details / progress / unknown"""
    result = ask_claude(
        "Classify this fitness WhatsApp message.\n\n"
        "Message: " + message + "\n\n"
        "Intents:\n"
        "- update: client updating weight, diet, height or fitness goal\n"
        "- my_details: client wants their saved profile file\n"
        "- progress: client wants their weight progress report\n"
        "- unknown: none of the above\n\n"
        "Reply ONE word only: update, my_details, progress, or unknown",
        max_tokens=10
    )
    print("Client intent: " + result)
    return result.lower()
 
def parse_registration(message):
    """Extract client details from registration message"""
    raw = ask_claude(
        "Extract fitness client details from this message.\n\n"
        "Message: " + message + "\n\n"
        "Extract: name, age, weight (with unit), height (with unit), blood_group, diet_type\n"
        "Use 'Not provided' for missing fields.\n\n"
        "Reply ONLY with JSON. No markdown. No backticks:\n"
        "{\"name\":\"Rahul Sharma\",\"age\":\"28\",\"weight\":\"82kg\","
        "\"height\":\"175cm\",\"blood_group\":\"B+\",\"diet_type\":\"Vegetarian\"}"
    )
    print("Registration: " + raw)
    try:
        return json.loads(raw)
    except:
        return None
 
def parse_client_phone(message):
    """Extract client phone number from trainer registration message"""
    raw = ask_claude(
        "Extract the client phone number from this message.\n\n"
        "Message: " + message + "\n\n"
        "It may appear as: phone 919533526497, mobile 919533526497, number 919533526497\n\n"
        "Reply ONLY with the digits (no spaces, no +, no dashes).\n"
        "If none found reply: none",
        max_tokens=30
    )
    raw = raw.strip().lower()
    print("Client phone: " + raw)
    return None if (raw == "none" or not raw.isdigit()) else raw
 
def parse_update(message):
    """Extract update fields from message.
    Known fields: age, weight, height, diet_type, fitness_goal
    Anything else goes into _comment field and saved as a note."""
    raw = ask_claude(
        "Extract what to update from this fitness message.\n\n"
        "Message: " + message + "\n\n"
        "Known fields: age, weight, height, diet_type, fitness_goal\n\n"
        "Rules:\n"
        "1. Extract any known fields with their values\n"
        "2. If message contains info that does not fit any known field, "
        "put it in a field called _comment\n"
        "3. If nothing found at all, return {}\n\n"
        "Reply ONLY with JSON. No markdown. No backticks.\n"
        "Examples:\n"
        "Input: update weight 79kg age 30 -> {\"age\":\"30\",\"weight\":\"79kg\"}\n"
        "Input: completed 30 pushups today -> {\"_comment\":\"completed 30 pushups today\"}\n"
        "Input: weight 79kg, started gym -> {\"weight\":\"79kg\",\"_comment\":\"started gym\"}\n"
        "If nothing found: {}"
    )
    try:
        return json.loads(raw)
    except:
        return {}
 
def parse_trainer_command(message):
    """Parse trainer command into action + parameters"""
    raw = ask_claude(
        "Parse this trainer command for a fitness bot.\n\n"
        "Command: " + message + "\n\n"
        "Extract:\n"
        "- action: register, update, note, progress, details, list\n"
        "- identifier: client ID (RAH-001) or name for update/note/progress/details\n"
        "- field: weight/height/diet_type/fitness_goal if action is update\n"
        "- value: new value if action is update\n"
        "- note_text: note content if action is note\n\n"
        "Reply ONLY with JSON. No markdown. No backticks:\n"
        "{\"action\":\"update\",\"identifier\":\"RAH-001\","
        "\"field\":\"weight\",\"value\":\"79kg\",\"note_text\":\"\"}"
    )
    print("Trainer parse: " + raw)
    try:
        return json.loads(raw)
    except:
        return None
 
# ══════════════════════════════════════════════════════
# EXCEL FUNCTIONS
# ══════════════════════════════════════════════════════
 
def create_excel(data, client_id):
    """Create a new styled Excel file for a client"""
    name     = data.get('name', 'Unknown')
    safe     = name.replace(' ','_').replace('/','_').replace('\\','_')
    filename = client_id.replace('-','') + "_" + safe + ".xlsx"
    filepath = os.path.join(DESKTOP_PATH, filename)
    now      = datetime.now().strftime("%d-%b-%Y %H:%M")
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FitTrack Profile"
 
    # Header row — blue background
    ws.append(HEADERS)
    hfill = PatternFill(start_color="1F5C99", end_color="1F5C99", fill_type="solid")
    for col in range(1, len(HEADERS)+1):
        c = ws.cell(row=1, column=col)
        c.fill = hfill
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    # Registration row — green background
    row = [
        "Registration", client_id, name,
        data.get('age','Not provided'), data.get('weight','Not provided'),
        data.get('height','Not provided'), data.get('blood_group','Not provided'),
        data.get('diet_type','Not provided'), "Not provided", "", now, "Trainer"
    ]
    ws.append(row)
    rfill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
    for col in range(1, len(row)+1):
        c = ws.cell(row=2, column=col)
        c.fill = rfill
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
 
    # Column widths
    for i, w in enumerate([16,12,20,8,10,10,14,18,20,25,20,14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
 
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    wb.save(filepath)
    print("Excel created: " + filepath)
    return filepath, filename
 
def add_row(filepath, updates, row_type, note="", updated_by="Client"):
    """Append a new row to existing Excel file"""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    last = [ws.cell(row=ws.max_row, column=c).value for c in range(1,13)]
    now  = datetime.now().strftime("%d-%b-%Y %H:%M")
 
    new_row = [
        row_type, last[1], last[2],
        updates.get('age',         last[3]),
        updates.get('weight',      last[4]),
        updates.get('height',      last[5]),
        last[6],
        updates.get('diet_type',   last[7]),
        updates.get('fitness_goal',last[8]),
        note if note else (last[9] or ""),
        now, updated_by
    ]
 
    # Row color by type
    colors = {
        "Client Update":  "FFF8E1",   # yellow
        "Client Note":    "FEF9E7",   # light yellow
        "Trainer Update": "EBF3FB",   # blue
        "Trainer Note":   "F5EEF8",   # purple
    }
    fill = PatternFill(
        start_color=colors.get(row_type,"FFFFFF"),
        end_color=  colors.get(row_type,"FFFFFF"),
        fill_type="solid"
    )
 
    ws.append(new_row)
    rn = ws.max_row
    for col in range(1, len(new_row)+1):
        c = ws.cell(row=rn, column=col)
        c.fill = fill
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rn].height = 20
    wb.save(filepath)
    print(row_type + " row added")
 
def build_progress_report(filepath, name, client_id):
    """Build a weight progress report from Excel data"""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[2]]
 
    if not rows:
        return "No data found in your profile."
 
    sw = str(rows[0][4]);  sd = str(rows[0][10])
    lw = str(rows[-1][4]); ld = str(rows[-1][10])
 
    try:
        s = float(''.join(c for c in sw if c.isdigit() or c=='.'))
        l = float(''.join(c for c in lw if c.isdigit() or c=='.'))
        d = round(s - l, 1)
        if d > 0:
            change = str(d)  + "kg lost";   status = "Amazing progress! Keep going!"
        elif d < 0:
            change = str(-d) + "kg gained"; status = "Stay focused on your goal!"
        else:
            change = "No change yet";        status = "Stay consistent — results take time!"
    except:
        change = "Could not calculate"; status = "Keep tracking!"
 
    history = "Start:   " + sw + " (" + sd + ")\n"
    for i, r in enumerate(rows[1:], 1):
        history += "Week " + str(i) + ":   " + str(r[4]) + " (" + str(r[10]) + ") [" + str(r[0]) + "]\n"
 
    return (
        "FitTrack Progress Report\n"
        + "-"*30 + "\n"
        + "Client: " + name + "\n"
        + "ID: "     + client_id + "\n\n"
        + "Weight History:\n" + history + "\n"
        + "Total Change: " + change + "\n"
        + "Check-ins: "   + str(len(rows)) + " entries\n\n"
        + status
    )
 
# ══════════════════════════════════════════════════════
# TRAINER — REGISTER CLIENT
# ══════════════════════════════════════════════════════
 
def trainer_register(body, trainer):
    send_text(trainer, "Registering client... please wait!")
 
    data = parse_registration(body)
    if not data or data.get('name') in [None, 'Unknown', 'Not provided']:
        send_text(trainer,
            "Could not extract client details. Try:\n\n"
            "trainer register Rahul Sharma, 28, 82kg, 175cm, B+, Vegetarian, phone 919533526497"
        )
        return
 
    client_phone = parse_client_phone(body)
    name         = data.get('name','Unknown')
    client_id    = generate_id(name)
    fp, fn       = create_excel(data, client_id)
 
    if client_phone:
        register_client(client_phone, client_id, name, fn)
 
    now     = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack — " + name + " (" + client_id + ") — " + now
 
    # ── Notify TRAINER ──
    send_text(trainer,
        "Client registered!\n\n"
        + "="*25 + "\n"
        + "Client ID: *" + client_id + "*\n"
        + "="*25 + "\n\n"
        + "Name:        " + name + "\n"
        + "Age:         " + data.get('age','N/A') + "\n"
        + "Weight:      " + data.get('weight','N/A') + "\n"
        + "Height:      " + data.get('height','N/A') + "\n"
        + "Blood Group: " + data.get('blood_group','N/A') + "\n"
        + "Diet:        " + data.get('diet_type','N/A') + "\n"
        + "Phone:       " + (client_phone or "Not provided") + "\n\n"
        + "Trainer commands:\n"
        + "trainer update " + client_id + " weight 79kg\n"
        + "trainer note "   + client_id + " great session today\n"
        + "trainer progress " + client_id
    )
    send_file(trainer, fp, fn, "New Client — " + caption)
 
    # ── Notify CLIENT ──
    if client_phone:
        send_text(client_phone,
            "Welcome to FitTrack, " + name + "!\n\n"
            + "="*25 + "\n"
            + "Your FitTrack ID: *" + client_id + "*\n"
            + "="*25 + "\n\n"
            + "Your trainer has registered you!\n\n"
            + "Your Profile:\n"
            + "Name:        " + name + "\n"
            + "Age:         " + data.get('age','N/A') + "\n"
            + "Weight:      " + data.get('weight','N/A') + "\n"
            + "Height:      " + data.get('height','N/A') + "\n"
            + "Blood Group: " + data.get('blood_group','N/A') + "\n"
            + "Diet Type:   " + data.get('diet_type','N/A') + "\n\n"
            + "Commands:\n"
            + "update weight 79kg\n"
            + "update goal muscle gain\n"
            + "my details\n"
            + "my progress"
        )
        send_file(client_phone, fp, fn, "Welcome — " + caption)
        print("Client notified: " + client_phone)
    else:
        send_text(trainer,
            "Note: No client phone found.\n"
            "Client NOT notified.\n"
            "Share their ID manually: *" + client_id + "*"
        )
    print("Registration done: " + name + " (" + client_id + ")")
 
# ══════════════════════════════════════════════════════
# TRAINER — UPDATE / NOTE
# ══════════════════════════════════════════════════════
 
def trainer_update(parsed, trainer):
    identifier = parsed.get('identifier','').strip()
    action     = parsed.get('action','').lower()
    field      = parsed.get('field','').strip()
    value      = parsed.get('value','').strip()
    note_text  = parsed.get('note_text','').strip()
 
    if not identifier:
        send_text(trainer,
            "Please specify client ID or name.\n"
            "Example: trainer update RAH-001 weight 79kg"
        )
        return
 
    fp, fn, cid, name = find_client(identifier)
    if not fp:
        send_text(trainer,
            "Client not found: " + identifier + "\n"
            "Use 'trainer list' to see all clients."
        )
        return
 
    now     = datetime.now().strftime("%d-%b-%Y %H:%M")
    cap_now = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack — " + name + " (" + cid + ") — " + cap_now
 
    # ── Note ──
    if action == "note":
        if not note_text:
            send_text(trainer, "Provide note text.\nExample: trainer note RAH-001 great session today")
            return
        add_row(fp, {}, "Trainer Note", note=note_text, updated_by="Trainer")
        send_text(trainer,
            "Note saved!\n"
            + "Client: " + name + " (" + cid + ")\n"
            + "Note: "   + note_text + "\n"
            + "Date: "   + now
        )
        return
 
    # ── Update ──
    updates = {}
    if field and value:
        updates[field] = value
    if not updates:
        send_text(trainer, "Specify what to update.\nExample: trainer update RAH-001 weight 79kg")
        return
 
    add_row(fp, updates, "Trainer Update", updated_by="Trainer")
    updated_lines = [k.replace('_',' ').title() + ": " + v for k, v in updates.items()]
 
    # Notify TRAINER
    send_text(trainer,
        "Update saved!\n"
        + "Client: " + name + " (" + cid + ")\n\n"
        + "\n".join(updated_lines) + "\n"
        + "Date: " + now
    )
    send_file(trainer, fp, fn, "Trainer Updated — " + caption)
 
    # Notify CLIENT
    cp = get_phone_for_client(cid)
    if cp:
        send_text(cp,
            "Your trainer updated your profile!\n\n"
            + "ID: " + cid + "\n"
            + "\n".join(updated_lines) + "\n"
            + "Date: " + now + "\n\n"
            + "Send 'my progress' to see your journey!"
        )
        send_file(cp, fp, fn, caption)
 
# ══════════════════════════════════════════════════════
# TRAINER — ALL COMMANDS ROUTER
# ══════════════════════════════════════════════════════
 
def handle_trainer(body, sender):
    parsed = parse_trainer_command(body)
    if not parsed:
        send_text(sender,
            "Could not understand command.\n\n"
            "Trainer commands:\n"
            "trainer register Rahul, 28, 82kg, 175cm, B+, Veg, phone 91XXXXXXXXXX\n"
            "trainer update RAH-001 weight 79kg\n"
            "trainer note RAH-001 great session today\n"
            "trainer progress RAH-001\n"
            "trainer details RAH-001\n"
            "trainer list"
        )
        return
 
    action = parsed.get('action','').lower()
    print("Trainer action: " + action)
 
    if action == "register":
        trainer_register(body, sender)
 
    elif action == "list":
        m   = load_mapping()
        ids = m.get("ids", {})
        if not ids:
            send_text(sender, "No clients registered yet.")
            return
        lines = ["FitTrack Client List", "-"*30, "Total: " + str(len(ids)) + " clients\n"]
        for cid, info in sorted(ids.items()):
            lines.append(
                cid + " — " + info["name"] + "\n"
                "     Phone: " + info.get("phone","N/A") + "\n"
                "     Joined: " + info.get("registered","N/A")
            )
        send_text(sender, "\n".join(lines))
 
    elif action == "progress":
        ident = parsed.get('identifier','').strip()
        if not ident:
            send_text(sender, "Specify client.\nExample: trainer progress RAH-001")
            return
        fp, fn, cid, name = find_client(ident)
        if not fp:
            send_text(sender, "Client not found: " + ident)
            return
        send_text(sender, build_progress_report(fp, name, cid))
 
    elif action == "details":
        ident = parsed.get('identifier','').strip()
        if not ident:
            send_text(sender, "Specify client.\nExample: trainer details RAH-001")
            return
        fp, fn, cid, name = find_client(ident)
        if not fp:
            send_text(sender, "Client not found: " + ident)
            return
        send_text(sender, "Here is " + name + " (" + cid + ") profile:")
        send_file(sender, fp, fn, "FitTrack — " + name + " (" + cid + ")")
 
    elif action in ["update", "note"]:
        trainer_update(parsed, sender)
 
    else:
        send_text(sender, "Unknown trainer command. Send 'trainer list' to see all clients.")
 
# ══════════════════════════════════════════════════════
# CLIENT — UPDATE
# ══════════════════════════════════════════════════════
 
def client_update(body, sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return
 
    send_text(sender, "Updating your profile " + cid + "... please wait!")
    updates = parse_update(body)
    if not updates:
        send_text(sender,
            "Could not understand your message. Try:\n"
            "update weight 79kg\n"
            "update age 30\n"
            "update goal muscle gain\n"
            "update diet keto\n\n"
            "Or just write anything — it will be saved as a comment!"
        )
        return
 
    # Separate known fields from comment
    comment   = updates.pop('_comment', '')
    known     = updates  # remaining known fields
 
    # Add update row for known fields
    if known:
        add_row(fp, known, "Client Update", updated_by="Client")
 
    # Add separate comment row if comment exists
    if comment:
        add_row(fp, {}, "Client Note", note=comment, updated_by="Client")
 
    now     = datetime.now().strftime("%d-%b-%Y %H:%M")
    cap_now = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack Updated — " + name + " (" + cid + ") — " + cap_now
 
    # Build reply lines
    lines = [k.replace('_',' ').title() + ": " + v for k, v in known.items()]
    if comment:
        lines.append("Note saved: " + comment)
 
    # Notify CLIENT
    send_text(sender,
        "Profile updated!\n"
        + "ID: " + cid + "\n\n"
        + "\n".join(lines) + "\n\n"
        + "Date: " + now + "\n\n"
        + "Send 'my progress' to see your journey!"
    )
    send_file(sender, fp, fn, caption)
 
    # Notify TRAINER
    send_to_trainer(
        "Client update received!\n\n"
        + "Client: " + name + " (" + cid + ")\n\n"
        + "\n".join(lines) + "\n"
        + "Date: " + now,
        fp, fn, "Client Update — " + caption
    )
 
# ══════════════════════════════════════════════════════
# CLIENT — MY DETAILS
# ══════════════════════════════════════════════════════
 
def client_details(sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return
    send_text(sender, "Here is your FitTrack profile (" + cid + "):")
    send_file(sender, fp, fn,
        "FitTrack Profile — " + name + " (" + cid + ") — " + datetime.now().strftime("%d %b %Y"))
 
# ══════════════════════════════════════════════════════
# CLIENT — MY PROGRESS
# ══════════════════════════════════════════════════════
 
def client_progress(sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return
    send_text(sender, build_progress_report(fp, name, cid))
 
# ══════════════════════════════════════════════════════
# WEBHOOK — MAIN ENTRY POINT
# ══════════════════════════════════════════════════════
 
@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        data = request.get_json(force=True, silent=True) or request.form.to_dict()
        print("\n" + "="*50)
        print("FITTRACK WEBHOOK")
        print("="*50)
 
        msg  = data.get('data', data)
        body = msg.get('body','')
        sender  = msg.get('from','')
        from_me = msg.get('fromMe', False)
 
        print("From: " + sender)
        print("Body: " + body)
 
        # ── Guard clauses ──
        if not body or not sender:
            return 'ok', 200
        if from_me is True or str(from_me).lower() == 'true':
            return 'ok', 200
        # Only block the UltraMsg bot number (OWN_NUMBER)
        # Do NOT block trainer personal numbers — they need to send commands!
        sender_clean = clean_number(sender)
        own_clean    = clean_number(OWN_NUMBER)
        if sender_clean == own_clean or own_clean in sender_clean or sender_clean in own_clean:
            print("Skipping own bot number: " + sender_clean)
            return 'ok', 200
 
        # ── TRAINER ──
        if body.strip().lower().startswith('trainer'):
            if is_trainer(sender):
                handle_trainer(body, sender)
            else:
                send_text(sender,
                    "You are not authorised as a trainer.\n\n"
                    "If you are a client, contact your trainer to register you on FitTrack."
                )
            return 'ok', 200
 
        # ── CLIENT — must be registered ──
        if not is_registered(sender):
            send_text(sender,
                "Welcome to FitTrack!\n\n"
                "You are not registered yet.\n\n"
                "Please contact your trainer to get registered.\n"
                "Your trainer will set up your profile and send you your unique FitTrack ID."
            )
            return 'ok', 200
 
        # ── REGISTERED CLIENT ──
        intent = detect_client_intent(body)
 
        if intent == "update":
            client_update(body, sender)
        elif intent == "my_details":
            client_details(sender)
        elif intent == "progress":
            client_progress(sender)
        else:
            send_text(sender,
                "FitTrack Commands:\n\n"
                "UPDATE your details:\n"
                "update weight 79kg\n"
                "update goal muscle gain\n"
                "update diet keto\n\n"
                "GET YOUR FILE:\n"
                "my details\n\n"
                "SEE YOUR PROGRESS:\n"
                "my progress"
            )
 
        print("="*50 + "\n")
        return 'ok', 200
 
    except Exception as e:
        print("ERROR: " + str(e))
        import traceback; traceback.print_exc()
        return 'error', 500
 
# ══════════════════════════════════════════════════════
# UTILITY ROUTES
# ══════════════════════════════════════════════════════
 
@app.route('/', methods=['GET'])
def home():
    return "FitTrack Bot is running!"
 
@app.route('/clients', methods=['GET'])
def list_clients():
    m   = load_mapping()
    ids = m.get("ids", {})
    out = "FitTrack (" + str(len(ids)) + " clients)\n\n"
    for cid, info in sorted(ids.items()):
        out += cid + " — " + info["name"] + " — " + info.get("registered","") + "\n"
    return out
 
# ══════════════════════════════════════════════════════
# START
# ══════════════════════════════════════════════════════
 
if __name__ == '__main__':
    print("="*50)
    print("FitTrack Bot — Final Clean Version")
    print("Registration : Trainer ONLY")
    print("Updates      : Trainer + Client")
    print("ID Format    : RAH-001, PRI-001, RAH-002")
    print("Folder       : " + DESKTOP_PATH)
    print("Webhook      : http://localhost:5000/webhook")
    print("Clients      : http://localhost:5000/clients")
    print("="*50)
    app.run(host='0.0.0.0', port=5000, debug=True)