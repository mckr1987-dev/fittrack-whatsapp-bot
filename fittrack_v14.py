# ── FITTRACK BOT — FINAL CLEAN VERSION ──
# Chenna Kesava Reddy
#
# HOW IT WORKS:
# ─────────────────────────────────────────────────────
# IDENTIFICATION:
#   1. Message from OWN_NUMBER         → ignore (bot's own message)
#   2. Message starts with "trainer"   → check TRAINER_NUMBERS → process trainer command
#   3. Sender in phone_mapping.json    → registered client → process client command
#   4. Sender not in mapping           → unregistered → tell them to contact trainer
#
# TRAINER COMMANDS (only from TRAINER_NUMBERS):
#   trainer register Rahul, 28, 82kg, 175cm, B+, Veg, phone 919533526497
#   trainer update RAH-001 weight 79kg
#   trainer note RAH-001 great session today
#   trainer progress RAH-001
#   trainer details RAH-001
#   trainer list
#
# CLIENT COMMANDS (only registered clients):
#   update weight 79kg
#   update goal muscle gain
#   my details
#   my progress
# ─────────────────────────────────────────────────────

from flask import Flask, request
import requests as req
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os, json, base64
from datetime import datetime

# Load environment variables from .env file if present
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed — use system env vars or defaults

# ══════════════════════════════════════════════════════
# CONFIG — update these values
# ══════════════════════════════════════════════════════
ULTRAMSG_INSTANCE = "instance182370"
ULTRAMSG_TOKEN    = os.getenv("ULTRAMSG_TOKEN", "YOUR_ULTRAMSG_TOKEN_HERE")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "YOUR_ANTHROPIC_API_KEY_HERE")
DESKTOP_PATH      = os.getenv("DESKTOP_PATH", r"C:\Users\mckre\OneDrive\Desktop\FitTrack_Clients")

# Your UltraMsg number — bot ignores messages from this number
OWN_NUMBER = os.getenv("OWN_NUMBER", "YOUR_OWN_NUMBER_HERE")

# Trainer phone numbers — only these can send trainer commands
TRAINER_NUMBERS = [
    n.strip() for n in
    os.getenv("TRAINER_NUMBERS", "918123009006,917022206001").split(",")
]

# Derived URLs — no need to change
ULTRAMSG_CHAT_URL = "https://api.ultramsg.com/" + ULTRAMSG_INSTANCE + "/messages/chat"

# ══════════════════════════════════════════════════════
# FIELD PERMISSIONS
# Defines what trainer and client can update
# ══════════════════════════════════════════════════════

# Fields ONLY trainer can update
TRAINER_ONLY_FIELDS = {
    'diet_type', 'diet', 'calories', 'calorie_intake',
    'payment_amount', 'payment_due_date', 'payment_due',
    'fee', 'monthly_fee', 'subscription_amount'
}

# Fields BOTH trainer and client can update
SHARED_FIELDS = {
    'weight', 'age', 'height', 'blood_sugar',
    'blood_pressure', 'blood_sugar_fasting',
    'blood_sugar_pp', 'resting_heart_rate',
    'chest', 'waist', 'hips', 'thighs', 'arms',
    'body_fat_percentage', 'bmi', 'fitness_goal',
    'steps_per_day', 'sleep_hours', 'water_intake'
}

# Payment status values — different for trainer vs client
TRAINER_PAYMENT_STATUS = "Confirmed by Trainer"
CLIENT_PAYMENT_STATUS  = "Paid by Client"
ULTRAMSG_FILE_URL = "https://api.ultramsg.com/" + ULTRAMSG_INSTANCE + "/messages/document"
MAPPING_FILE      = os.path.join(DESKTOP_PATH, "phone_mapping.json")

# Note: Excel columns are fully dynamic — see FIXED_COLS_START and FIXED_COLS_END

# ══════════════════════════════════════════════════════
# APP INIT
# ══════════════════════════════════════════════════════
app = Flask(__name__)

if not os.path.exists(DESKTOP_PATH):
    os.makedirs(DESKTOP_PATH)
    print("Created folder: " + DESKTOP_PATH)

# ══════════════════════════════════════════════════════
# MAPPING FUNCTIONS
# phone_mapping.json stores:
#   phones  : { "919533526497": "RAH-001" }
#   ids     : { "RAH-001": { name, filename, phone, registered } }
#   clients : { "rahul sharma": "RAH-001" }
#   prefixes: { "RAH": 1, "PRI": 2 }
# ══════════════════════════════════════════════════════

def load_mapping():
    if os.path.exists(MAPPING_FILE):
        with open(MAPPING_FILE, 'r') as f:
            return json.load(f)
    return {"phones": {}, "ids": {}, "clients": {}, "prefixes": {}}

def save_mapping(data):
    with open(MAPPING_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def clean_number(phone):
    """Strip WhatsApp suffixes and normalize to include country code.
    Ensures 919916696004 and 9916696004 both resolve to 919916696004
    so trainer-registered numbers always match client incoming numbers."""
    # Strip WhatsApp suffixes
    number = phone.replace('@c.us','').replace('@lid','').split('@')[0].strip()
    # Remove any leading + sign
    number = number.lstrip('+')
    # Normalize Indian numbers — if 10 digits starting with 6-9, add 91
    if len(number) == 10 and number[0] in '6789':
        number = '91' + number
    return number

def generate_id(name):
    """RAH-001 for Rahul, PRI-001 for Priya, RAH-002 for second Rahul"""
    m      = load_mapping()
    prefix = name.strip().split()[0][:3].upper()
    count  = m.get("prefixes", {}).get(prefix, 0) + 1
    m.setdefault("prefixes", {})[prefix] = count
    save_mapping(m)
    return prefix + "-" + str(count).zfill(3)

def register_client(phone, client_id, name, filename):
    """Save phone → client_id mapping after registration"""
    m  = load_mapping()
    cp = clean_number(phone)
    m["phones"][cp]       = client_id
    m["clients"][name.lower()] = client_id
    m["ids"][client_id]   = {
        "name": name, "filename": filename,
        "phone": cp,
        "registered": datetime.now().strftime("%d-%b-%Y %H:%M")
    }
    save_mapping(m)
    print("Registered: " + cp + " → " + client_id)

def get_client_by_phone(sender):
    """Find client record using sender's phone number"""
    m   = load_mapping()
    cp  = clean_number(sender)
    cid = m.get("phones", {}).get(cp)
    if not cid:
        return None, None, None, None
    info = m.get("ids", {}).get(cid, {})
    fp   = os.path.join(DESKTOP_PATH, info.get("filename", ""))
    if os.path.exists(fp):
        return fp, info["filename"], cid, info["name"]
    return None, None, None, None

def get_client_by_id(client_id):
    """Find client record using client ID like RAH-001"""
    m    = load_mapping()
    cid  = client_id.upper().strip()
    info = m.get("ids", {}).get(cid, {})
    if not info:
        return None, None, None, None
    fp = os.path.join(DESKTOP_PATH, info.get("filename", ""))
    if os.path.exists(fp):
        return fp, info["filename"], cid, info["name"]
    return None, None, None, None

def get_client_by_name(name):
    """Find client record using client name (partial match supported)"""
    m  = load_mapping()
    nl = name.lower().strip()
    # exact match
    cid = m.get("clients", {}).get(nl)
    if cid:
        return get_client_by_id(cid)
    # partial match
    for stored_name, cid in m.get("clients", {}).items():
        if nl in stored_name or stored_name in nl:
            return get_client_by_id(cid)
    return None, None, None, None

def find_client(identifier):
    """Smart lookup — works with ID (RAH-001) or name (Rahul)"""
    identifier = identifier.strip()
    parts = identifier.upper().split('-')
    if len(parts) == 2 and parts[0].isalpha() and parts[1].isdigit():
        return get_client_by_id(identifier)
    return get_client_by_name(identifier)

def get_phone_for_client(client_id):
    """Get client's phone number from mapping"""
    m = load_mapping()
    return m.get("ids", {}).get(client_id.upper(), {}).get("phone", "")

def trainer_phone():
    """Return first non-UltraMsg trainer number"""
    for t in TRAINER_NUMBERS:
        if t != OWN_NUMBER:
            return t
    return TRAINER_NUMBERS[0]

def is_trainer(sender):
    """Check if sender is a registered trainer"""
    cp = clean_number(sender)
    return any(t in cp or cp in t for t in TRAINER_NUMBERS)

def is_registered(sender):
    """Check if sender is a registered client"""
    fp, *_ = get_client_by_phone(sender)
    return fp is not None

# ══════════════════════════════════════════════════════
# WHATSAPP FUNCTIONS
# ══════════════════════════════════════════════════════

def send_text(to, message):
    """Send a WhatsApp text message"""
    cp = clean_number(to)
    try:
        r = req.post(ULTRAMSG_CHAT_URL, timeout=10,
                     data={"token": ULTRAMSG_TOKEN, "to": cp, "body": message})
        print("Text → " + cp + " [" + str(r.status_code) + "]")
    except Exception as e:
        print("Text error: " + str(e))

def send_file(to, filepath, filename, caption):
    """Send a WhatsApp file attachment"""
    cp = clean_number(to)
    try:
        with open(filepath, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('utf-8')
        mime = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        r = req.post(ULTRAMSG_FILE_URL, timeout=15, data={
            "token": ULTRAMSG_TOKEN, "to": cp,
            "filename": filename, "caption": caption,
            "document": mime + b64
        })
        print("File → " + cp + " [" + str(r.status_code) + "]")
    except Exception as e:
        print("File error: " + str(e))

def send_to_trainer(message, fp=None, fn=None, caption=None):
    """Send message (and optionally file) to trainer"""
    send_text(trainer_phone(), message)
    if fp and fn and caption:
        send_file(trainer_phone(), fp, fn, caption)

def send_to_client(client_id, message, fp=None, fn=None, caption=None):
    """Send message (and optionally file) to client using their ID"""
    cp = get_phone_for_client(client_id)
    if not cp:
        print("No phone found for: " + client_id)
        return
    send_text(cp, message)
    if fp and fn and caption:
        send_file(cp, fp, fn, caption)

# ══════════════════════════════════════════════════════
# CLAUDE AI FUNCTIONS
# ══════════════════════════════════════════════════════

def ask_claude(prompt, max_tokens=300):
    """Send a prompt to Claude and return clean text"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    r = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}]
    )
    return r.content[0].text.strip().replace('```json','').replace('```','').strip()

def detect_client_intent(message):
    """Classify client message as update / my_details / progress / query / unknown"""
    result = ask_claude(
        "Classify this fitness WhatsApp message.\n\n"
        "Message: " + message + "\n\n"
        "Intents:\n"
        "- update: client updating weight, diet, height or fitness goal\n"
        "- my_details: client wants their saved profile file\n"
        "- progress: client wants their weight progress report\n"
        "- query: client is asking a question about their data, history, stats or progress\n"
        "- unknown: none of the above\n\n"
        "Reply ONE word only: update, my_details, progress, query, or unknown",
        max_tokens=10
    )
    print("Client intent: " + result)
    return result.lower()

def parse_registration(message):
    """Extract client details from registration message"""
    raw = ask_claude(
        "Extract fitness client details from this message.\n\n"
        "Message: " + message + "\n\n"
        "Extract: name, age, weight (with unit), height (with unit), blood_group, diet_type\n"
        "Use 'Not provided' for missing fields.\n\n"
        "Reply ONLY with JSON. No markdown. No backticks:\n"
        "{\"name\":\"Rahul Sharma\",\"age\":\"28\",\"weight\":\"82kg\","
        "\"height\":\"175cm\",\"blood_group\":\"B+\",\"diet_type\":\"Vegetarian\"}"
    )
    print("Registration: " + raw)
    try:
        return json.loads(raw)
    except:
        return None

def parse_client_phone(message):
    """Extract client phone number from trainer registration message"""
    raw = ask_claude(
        "Extract the client phone number from this message.\n\n"
        "Message: " + message + "\n\n"
        "It may appear as: phone 919533526497, mobile 919533526497, number 919533526497\n\n"
        "Reply ONLY with the digits (no spaces, no +, no dashes).\n"
        "If none found reply: none",
        max_tokens=30
    )
    raw = raw.strip().lower()
    print("Client phone: " + raw)
    return None if (raw == "none" or not raw.isdigit()) else raw

def parse_update(message):
    """Extract update fields from message.
    Known fields: age, weight, height, diet_type, fitness_goal
    Anything else goes into _comment field and saved as a note."""
    raw = ask_claude(
        "Extract what to update from this fitness message.\n\n"
        "Message: " + message + "\n\n"
        "Known fields: age, weight, height, diet_type, fitness_goal\n\n"
        "Rules:\n"
        "1. Extract any known fields with their values\n"
        "2. If message contains info that does not fit any known field, "
        "put it in a field called _comment\n"
        "3. If nothing found at all, return {}\n\n"
        "Reply ONLY with JSON. No markdown. No backticks.\n"
        "Examples:\n"
        "Input: update weight 79kg age 30 -> {\"age\":\"30\",\"weight\":\"79kg\"}\n"
        "Input: completed 30 pushups today -> {\"_comment\":\"completed 30 pushups today\"}\n"
        "Input: weight 79kg, started gym -> {\"weight\":\"79kg\",\"_comment\":\"started gym\"}\n"
        "If nothing found: {}"
    )
    try:
        return json.loads(raw)
    except:
        return {}

def parse_bulk_history(message):
    """Detect and extract bulk weight history from message.
    Returns list of {weight, date/week} dicts if bulk history detected.
    Returns None if not bulk history."""
    raw = ask_claude(
        "Analyse this fitness message and check if it contains MULTIPLE weight entries.\n\n"
        "Message: " + message + "\n\n"
        "Rules:\n"
        "1. If message contains multiple weight values (more than 1) with week numbers or dates,\n"
        "   extract each as a separate entry\n"
        "2. Each entry should have: weight and week/date label\n"
        "3. If only ONE weight value mentioned, return: null\n"
        "4. If no weights at all, return: null\n\n"
        "Reply ONLY with JSON array or null. No markdown. No backticks.\n"
        "Example for: Week 1: 79kg, Week 2: 78kg, Week 3: 77.5kg\n"
        "[\n"
        "  {\"week\": \"Week 1\", \"weight\": \"79kg\"},\n"
        "  {\"week\": \"Week 2\", \"weight\": \"78kg\"},\n"
        "  {\"week\": \"Week 3\", \"weight\": \"77.5kg\"}\n"
        "]\n\n"
        "Example for: 79kg, 78kg, 77.5kg (no week labels):\n"
        "[\n"
        "  {\"week\": \"Week 1\", \"weight\": \"79kg\"},\n"
        "  {\"week\": \"Week 2\", \"weight\": \"78kg\"},\n"
        "  {\"week\": \"Week 3\", \"weight\": \"77.5kg\"}\n"
        "]\n\n"
        "For single weight like: update weight 79kg\n"
        "Return: null"
    )
    raw = raw.strip()
    if raw.lower() == 'null' or raw == '':
        return None
    try:
        result = json.loads(raw)
        if isinstance(result, list) and len(result) > 1:
            return result
        return None
    except:
        return None

def parse_trainer_command(message):
    """Parse trainer command into action + parameters"""
    raw = ask_claude(
        "Parse this trainer command for a fitness bot.\n\n"
        "Command: " + message + "\n\n"
        "Extract:\n"
        "- action: register, update, note, progress, details, list, query\n"
        "- identifier: client ID (RAH-001) or name for update/note/progress/details\n"
        "- field: weight/height/diet_type/fitness_goal if action is update\n"
        "- value: new value if action is update\n"
        "- note_text: note content if action is note\n\n"
        "Reply ONLY with JSON. No markdown. No backticks:\n"
        "{\"action\":\"update\",\"identifier\":\"RAH-001\","
        "\"field\":\"weight\",\"value\":\"79kg\",\"note_text\":\"\"}"
    )
    print("Trainer parse: " + raw)
    try:
        return json.loads(raw)
    except:
        return None

# ══════════════════════════════════════════════════════
# EXCEL FUNCTIONS
# ══════════════════════════════════════════════════════

# Fixed columns always present in every Excel file
FIXED_COLS_START = ["Type", "Client ID", "Client Name"]
FIXED_COLS_END   = ["Notes", "Date", "Updated By"]

def format_col_name(key):
    """Convert snake_case key to readable Title Case column name"""
    return key.replace('_', ' ').title()

def create_excel(data, client_id):
    """Create a new styled Excel file with DYNAMIC columns based on available data.
    Fixed columns: Type, Client ID, Client Name ... Notes, Date, Updated By
    Dynamic columns: whatever Claude extracted from the message"""
    name     = data.get('name', 'Unknown')
    safe     = name.replace(' ','_').replace('/','_').replace('\\','_')
    filename = client_id.replace('-','') + "_" + safe + ".xlsx"
    filepath = os.path.join(DESKTOP_PATH, filename)
    # Use custom registration date if provided, else use today
    custom_date = data.get('_registration_date', '')
    if custom_date:
        now = custom_date
        print("Using custom registration date: " + now)
    else:
        now = datetime.now().strftime("%d-%b-%Y %H:%M")

    # Build dynamic columns from data keys
    # Exclude fixed fields and internal fields
    skip_keys = {'name', '_comment', '_registration_date'}
    dynamic_keys = [k for k in data.keys() if k not in skip_keys]

    # Build full header list
    headers = FIXED_COLS_START + [format_col_name(k) for k in dynamic_keys] + FIXED_COLS_END

    # Build data row
    notes = data.get('_comment', '')
    row   = (
        ["Registration", client_id, name] +
        [str(data.get(k, '')) for k in dynamic_keys] +
        [notes, now, "Trainer"]
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FitTrack Profile"

    # Header row — teal background
    ws.append(headers)
    hfill = PatternFill(start_color="028090", end_color="028090", fill_type="solid")
    for col in range(1, len(headers)+1):
        c = ws.cell(row=1, column=col)
        c.fill = hfill
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Registration row — green background
    ws.append(row)
    rfill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
    for col in range(1, len(row)+1):
        c = ws.cell(row=2, column=col)
        c.fill = rfill
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Auto column widths
    for col in range(1, len(headers)+1):
        letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[letter].width = max(14, len(str(headers[col-1])) + 4)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    wb.save(filepath)
    print("Dynamic Excel created: " + filepath)
    print("Columns: " + str(headers))
    return filepath, filename

def add_row(filepath, updates, row_type, note="", updated_by="Client"):
    """Append a new dynamic row to existing Excel file.
    Reads current headers from Excel and maps update values to correct columns.
    New columns are added automatically if update contains new fields."""
    wb  = openpyxl.load_workbook(filepath)
    ws  = wb.active
    now = datetime.now().strftime("%d-%b-%Y %H:%M")

    # Read current headers from row 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

    # Read last data row to carry forward unchanged values
    last_row_num = ws.max_row
    last = {headers[c-1]: ws.cell(row=last_row_num, column=c).value for c in range(1, ws.max_column+1)}

    # Check if updates contain NEW columns not in headers
    skip_keys   = {'_comment'}
    update_keys = [k for k in updates.keys() if k not in skip_keys]

    for key in update_keys:
        col_name = format_col_name(key)
        if col_name not in headers:
            # Add new column to Excel!
            new_col_idx = headers.index("Notes") + 1  # insert before Notes
            ws.insert_cols(new_col_idx)

            # Add header for new column
            header_cell = ws.cell(row=1, column=new_col_idx, value=col_name)
            header_cell.fill = PatternFill(start_color="028090", end_color="028090", fill_type="solid")
            header_cell.font = Font(bold=True, color="FFFFFF", size=12)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(new_col_idx)].width = max(14, len(col_name)+4)

            # Refresh headers list
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            last    = {headers[c-1]: ws.cell(row=last_row_num, column=c).value for c in range(1, ws.max_column+1)}
            print("New column added: " + col_name)

    # Build new row using current headers
    new_row = []
    for h in headers:
        if h == "Type":
            new_row.append(row_type)
        elif h == "Date":
            new_row.append(now)
        elif h == "Updated By":
            new_row.append(updated_by)
        elif h == "Notes":
            new_row.append(note if note else (last.get("Notes") or ""))
        else:
            # Check if this column is being updated
            key = h.lower().replace(' ', '_')
            if key in updates:
                new_row.append(str(updates[key]))
            else:
                # Carry forward last value
                new_row.append(last.get(h, ""))

    # Row color by type
    colors = {
        "Client Update":  "FFF8E1",
        "Client Note":    "FEF9E7",
        "Trainer Update": "EBF3FB",
        "Trainer Note":   "F5EEF8",
    }
    fill = PatternFill(
        start_color=colors.get(row_type, "FFFFFF"),
        end_color=  colors.get(row_type, "FFFFFF"),
        fill_type="solid"
    )

    ws.append(new_row)
    rn = ws.max_row
    for col in range(1, len(new_row)+1):
        c = ws.cell(row=rn, column=col)
        c.fill = fill
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rn].height = 20
    wb.save(filepath)
    print(row_type + " row added dynamically")

def build_progress_report(filepath, name, client_id):
    """Build a weight progress report from Excel data.
    Uses dynamic column lookup instead of hardcoded indices."""
    wb   = openpyxl.load_workbook(filepath)
    ws   = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[2]]

    if not rows:
        return "No data found in your profile."

    # Dynamic column index lookup
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    weight_idx = next((i for i, h in enumerate(headers) if h and 'weight' in str(h).lower()), 4)
    date_idx   = next((i for i, h in enumerate(headers) if h and h == 'Date'), 10)

    sw = str(rows[0][weight_idx]);  sd = str(rows[0][date_idx])
    lw = str(rows[-1][weight_idx]); ld = str(rows[-1][date_idx])

    try:
        s = float(''.join(c for c in sw if c.isdigit() or c=='.'))
        l = float(''.join(c for c in lw if c.isdigit() or c=='.'))
        d = round(s - l, 1)
        if d > 0:
            change = str(d)  + "kg lost";   status = "Amazing progress! Keep going!"
        elif d < 0:
            change = str(-d) + "kg gained"; status = "Stay focused on your goal!"
        else:
            change = "No change yet";        status = "Stay consistent — results take time!"
    except:
        change = "Could not calculate"; status = "Keep tracking!"

    history = "Start:   " + sw + " (" + sd + ")\n"
    for i, r in enumerate(rows[1:], 1):
        history += "Week " + str(i) + ":   " + str(r[weight_idx]) + " (" + str(r[date_idx]) + ") [" + str(r[0]) + "]\n"

    return (
        "FitTrack Progress Report\n"
        + "-"*30 + "\n"
        + "Client: " + name + "\n"
        + "ID: "     + client_id + "\n\n"
        + "Weight History:\n" + history + "\n"
        + "Total Change: " + change + "\n"
        + "Check-ins: "   + str(len(rows)) + " entries\n\n"
        + status
    )

# ══════════════════════════════════════════════════════
# TRAINER — REGISTER CLIENT
# ══════════════════════════════════════════════════════

def trainer_register(body, trainer):
    send_text(trainer, "Registering client... please wait!")

    data = parse_registration(body)
    if not data or data.get('name') in [None, 'Unknown', 'Not provided']:
        send_text(trainer,
            "Could not extract client details. Try:\n\n"
            "trainer register Rahul Sharma, 28, 82kg, 175cm, B+, Vegetarian, phone 919533526497"
        )
        return

    client_phone = parse_client_phone(body)
    name         = data.get('name','Unknown')
    client_id    = generate_id(name)
    fp, fn       = create_excel(data, client_id)

    if client_phone:
        register_client(client_phone, client_id, name, fn)

    now     = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack — " + name + " (" + client_id + ") — " + now

    # ── Notify TRAINER ──
    send_text(trainer,
        "Client registered!\n\n"
        + "="*25 + "\n"
        + "Client ID: *" + client_id + "*\n"
        + "="*25 + "\n\n"
        + "Name:        " + name + "\n"
        + "Age:         " + data.get('age','N/A') + "\n"
        + "Weight:      " + data.get('weight','N/A') + "\n"
        + "Height:      " + data.get('height','N/A') + "\n"
        + "Blood Group: " + data.get('blood_group','N/A') + "\n"
        + "Diet:        " + data.get('diet_type','N/A') + "\n"
        + "Phone:       " + (client_phone or "Not provided") + "\n\n"
        + "Trainer commands:\n"
        + "trainer update " + client_id + " weight 79kg\n"
        + "trainer note "   + client_id + " great session today\n"
        + "trainer progress " + client_id
    )
    send_file(trainer, fp, fn, "New Client — " + caption)

    # ── Notify CLIENT ──
    if client_phone:
        send_text(client_phone,
            "Welcome to FitTrack, " + name + "!\n\n"
            + "="*25 + "\n"
            + "Your FitTrack ID: *" + client_id + "*\n"
            + "="*25 + "\n\n"
            + "Your trainer has registered you!\n\n"
            + "Your Profile:\n"
            + "Name:        " + name + "\n"
            + "Age:         " + data.get('age','N/A') + "\n"
            + "Weight:      " + data.get('weight','N/A') + "\n"
            + "Height:      " + data.get('height','N/A') + "\n"
            + "Blood Group: " + data.get('blood_group','N/A') + "\n"
            + "Diet Type:   " + data.get('diet_type','N/A') + "\n\n"
            + "Commands:\n"
            + "update weight 79kg\n"
            + "update goal muscle gain\n"
            + "my details\n"
            + "my progress"
        )
        send_file(client_phone, fp, fn, "Welcome — " + caption)
        print("Client notified: " + client_phone)
    else:
        send_text(trainer,
            "Note: No client phone found.\n"
            "Client NOT notified.\n"
            "Share their ID manually: *" + client_id + "*"
        )
    print("Registration done: " + name + " (" + client_id + ")")

# ══════════════════════════════════════════════════════
# TRAINER — UPDATE / NOTE
# ══════════════════════════════════════════════════════

def trainer_update(parsed, trainer):
    identifier = parsed.get('identifier','').strip()
    action     = parsed.get('action','').lower()
    field      = parsed.get('field','').strip()
    value      = parsed.get('value','').strip()
    note_text  = parsed.get('note_text','').strip()

    if not identifier:
        send_text(trainer,
            "Please specify client ID or name.\n"
            "Example: trainer update RAH-001 weight 79kg"
        )
        return

    fp, fn, cid, name = find_client(identifier)
    if not fp:
        send_text(trainer,
            "Client not found: " + identifier + "\n"
            "Use 'trainer list' to see all clients."
        )
        return

    now     = datetime.now().strftime("%d-%b-%Y %H:%M")
    cap_now = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack — " + name + " (" + cid + ") — " + cap_now

    # ── Note ──
    if action == "note":
        if not note_text:
            send_text(trainer, "Provide note text.\nExample: trainer note RAH-001 great session today")
            return
        add_row(fp, {}, "Trainer Note", note=note_text, updated_by="Trainer")
        send_text(trainer,
            "Note saved!\n"
            + "Client: " + name + " (" + cid + ")\n"
            + "Note: "   + note_text + "\n"
            + "Date: "   + now
        )
        return

    # ── Update ──
    updates = {}
    if field and value:
        updates[field] = value
    if not updates:
        send_text(trainer, "Specify what to update.\nExample: trainer update RAH-001 weight 79kg")
        return

    add_row(fp, updates, "Trainer Update", updated_by="Trainer")
    updated_lines = [k.replace('_',' ').title() + ": " + v for k, v in updates.items()]

    # Notify TRAINER
    send_text(trainer,
        "Update saved!\n"
        + "Client: " + name + " (" + cid + ")\n\n"
        + "\n".join(updated_lines) + "\n"
        + "Date: " + now
    )
    send_file(trainer, fp, fn, "Trainer Updated — " + caption)

    # Notify CLIENT
    cp = get_phone_for_client(cid)
    if cp:
        send_text(cp,
            "Your trainer updated your profile!\n\n"
            + "ID: " + cid + "\n"
            + "\n".join(updated_lines) + "\n"
            + "Date: " + now + "\n\n"
            + "Send 'my progress' to see your journey!"
        )
        send_file(cp, fp, fn, caption)

# ══════════════════════════════════════════════════════
# TRAINER — ALL COMMANDS ROUTER
# ══════════════════════════════════════════════════════

def handle_trainer(body, sender):
    parsed = parse_trainer_command(body)
    if not parsed:
        send_text(sender,
            "Could not understand command.\n\n"
            "Trainer commands:\n"
            "trainer register Rahul, 28, 82kg, 175cm, B+, Veg, phone 91XXXXXXXXXX\n"
            "trainer update RAH-001 weight 79kg\n"
            "trainer note RAH-001 great session today\n"
            "trainer progress RAH-001\n"
            "trainer details RAH-001\n"
            "trainer list"
        )
        return

    action = parsed.get('action','').lower()
    print("Trainer action: " + action)

    if action == "register":
        trainer_register(body, sender)

    elif action == "list":
        m   = load_mapping()
        ids = m.get("ids", {})
        if not ids:
            send_text(sender, "No clients registered yet.")
            return
        lines = ["FitTrack Client List", "-"*30, "Total: " + str(len(ids)) + " clients\n"]
        for cid, info in sorted(ids.items()):
            lines.append(
                cid + " — " + info["name"] + "\n"
                "     Phone: " + info.get("phone","N/A") + "\n"
                "     Joined: " + info.get("registered","N/A")
            )
        send_text(sender, "\n".join(lines))

    elif action == "progress":
        ident = parsed.get('identifier','').strip()
        if not ident:
            send_text(sender, "Specify client.\nExample: trainer progress RAH-001")
            return
        fp, fn, cid, name = find_client(ident)
        if not fp:
            send_text(sender, "Client not found: " + ident)
            return
        send_text(sender, build_progress_report(fp, name, cid))

    elif action == "details":
        ident = parsed.get('identifier','').strip()
        if not ident:
            send_text(sender, "Specify client.\nExample: trainer details RAH-001")
            return
        fp, fn, cid, name = find_client(ident)
        if not fp:
            send_text(sender, "Client not found: " + ident)
            return
        send_text(sender, "Here is " + name + " (" + cid + ") profile:")
        send_file(sender, fp, fn, "FitTrack — " + name + " (" + cid + ")")

    elif action in ["update", "note"]:
        trainer_update(parsed, sender)

    elif action == "query":
        handle_trainer_query(body, sender)

    else:
        send_text(sender, "Unknown trainer command. Send 'trainer list' to see all clients.")

# ══════════════════════════════════════════════════════
# CLIENT — UPDATE
# ══════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════
# FIELD VALIDATION HELPERS
# ══════════════════════════════════════════════════════

def normalise_field_key(key):
    """Normalise field key to standard snake_case for permission checking"""
    return key.lower().strip().replace(' ', '_').replace('-', '_')

def is_trainer_only(key):
    """Check if a field is restricted to trainer only"""
    nk = normalise_field_key(key)
    # Direct match
    if nk in TRAINER_ONLY_FIELDS:
        return True
    # Partial match — catches variations
    for restricted in TRAINER_ONLY_FIELDS:
        if restricted in nk or nk in restricted:
            return True
    return False

def is_payment_status(key):
    """Check if field is payment status"""
    nk = normalise_field_key(key)
    return 'payment' in nk and 'status' in nk

def validate_client_update(updates):
    """Split update fields into allowed, blocked and payment status.
    Returns: allowed_fields, blocked_fields, payment_status_value"""
    allowed  = {}
    blocked  = []
    pay_stat = None

    for key, value in updates.items():
        if key.startswith('_'):
            # Internal fields — skip
            continue
        elif is_payment_status(key):
            # Payment status — client can only say "Paid"
            pay_stat = CLIENT_PAYMENT_STATUS
        elif is_trainer_only(key):
            # Blocked for client
            blocked.append(key.replace('_', ' ').title())
        else:
            # Allowed for client
            allowed[key] = value

    return allowed, blocked, pay_stat

def check_payment_due(filepath, client_id, name, client_phone):
    """Check if payment is due and send reminder to client"""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Read headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

        # Find payment due date column
        due_col = None
        for i, h in enumerate(headers):
            if h and 'due' in str(h).lower() and 'payment' in str(h).lower():
                due_col = i + 1
                break

        if not due_col:
            return  # No payment due date column

        # Get latest due date value
        last_row = ws.max_row
        due_date = ws.cell(row=last_row, column=due_col).value

        if not due_date or due_date == '' or due_date == 'Not provided':
            return

        # Parse due date
        from datetime import datetime as dt
        today = dt.now()
        for fmt in ["%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]:
            try:
                due = dt.strptime(str(due_date), fmt)
                days_left = (due - today).days

                if days_left < 0:
                    # Overdue
                    send_text(client_phone,
                        "Payment Reminder 🔔\n\n"
                        "Hi " + name + "! Your FitTrack payment was due on " + str(due_date) + "\n"
                        "Your payment is " + str(abs(days_left)) + " days overdue.\n\n"
                        "Please contact your trainer to confirm payment."
                    )
                    print("Payment overdue reminder sent to: " + name)
                elif days_left <= 3:
                    # Due soon
                    send_text(client_phone,
                        "Payment Reminder 🔔\n\n"
                        "Hi " + name + "! Your FitTrack payment is due in " + str(days_left) + " day(s).\n"
                        "Due date: " + str(due_date) + "\n\n"
                        "Please arrange payment with your trainer."
                    )
                    print("Payment due soon reminder sent to: " + name)
                break
            except:
                continue

    except Exception as e:
        print("Payment check error: " + str(e))

def client_update(body, sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return

    send_text(sender, "Updating your profile " + cid + "... please wait!")

    # ── Check for bulk weight history first ──
    bulk = parse_bulk_history(body)
    if bulk:
        # Multiple weight entries detected — create one row per entry
        send_text(sender,
            "Bulk weight history detected!\n"
            "Creating " + str(len(bulk)) + " rows — one per week...\n\n"
            "Please wait!"
        )
        for entry in bulk:
            week   = entry.get('week', '')
            weight = entry.get('weight', '')
            if weight:
                note = week if week else ""
                add_row(fp, {'weight': weight}, "Client Update",
                        note=note, updated_by="Client")

        now     = datetime.now().strftime("%d-%b-%Y %H:%M")
        cap_now = datetime.now().strftime("%d %b %Y")
        caption = "FitTrack Updated — " + name + " (" + cid + ") — " + cap_now

        # Build summary
        summary = "Weight history added:\n"
        for entry in bulk:
            summary += "• " + entry.get('week','') + ": " + entry.get('weight','') + "\n"

        send_text(sender,
            "Done! " + str(len(bulk)) + " weight entries saved!\n\n"
            + summary + "\n"
            + "Send 'my progress' to see your full journey!"
        )
        send_file(sender, fp, fn, caption)

        # Notify trainer
        send_to_trainer(
            "Client bulk history received!\n\n"
            + "Client: " + name + " (" + cid + ")\n"
            + str(len(bulk)) + " weight entries added\n\n"
            + summary,
            fp, fn, "Bulk Update — " + caption
        )
        return

    # ── Single update ──
    updates = parse_update(body)
    if not updates:
        send_text(sender,
            "Could not understand your message. Try:\n"
            "update weight 79kg\n"
            "update age 30\n"
            "update goal muscle gain\n\n"
            "For bulk history send like:\n"
            "Week 1: 79kg, Week 2: 78kg, Week 3: 77.5kg"
        )
        return

    # Separate known fields from comment
    comment = updates.pop('_comment', '')
    known   = updates

    if known:
        add_row(fp, known, "Client Update", updated_by="Client")
    if comment:
        add_row(fp, {}, "Client Note", note=comment, updated_by="Client")

    now     = datetime.now().strftime("%d-%b-%Y %H:%M")
    cap_now = datetime.now().strftime("%d %b %Y")
    caption = "FitTrack Updated — " + name + " (" + cid + ") — " + cap_now
    lines   = [k.replace('_',' ').title() + ": " + v for k, v in known.items()]
    if comment:
        lines.append("Note saved: " + comment)

    send_text(sender,
        "Profile updated!\n"
        + "ID: " + cid + "\n\n"
        + "\n".join(lines) + "\n\n"
        + "Date: " + now + "\n\n"
        + "Send 'my progress' to see your journey!"
    )
    send_file(sender, fp, fn, caption)

    send_to_trainer(
        "Client update received!\n\n"
        + "Client: " + name + " (" + cid + ")\n\n"
        + "\n".join(lines) + "\n"
        + "Date: " + now,
        fp, fn, "Client Update — " + caption
    )

# ══════════════════════════════════════════════════════
# CLIENT — MY DETAILS
# ══════════════════════════════════════════════════════

def client_details(sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return
    send_text(sender, "Here is your FitTrack profile (" + cid + "):")
    send_file(sender, fp, fn,
        "FitTrack Profile — " + name + " (" + cid + ") — " + datetime.now().strftime("%d %b %Y"))

# ══════════════════════════════════════════════════════
# CLIENT — MY PROGRESS
# ══════════════════════════════════════════════════════

def client_progress(sender):
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return
    send_text(sender, build_progress_report(fp, name, cid))


# ══════════════════════════════════════════════════════
# MODULE 6: INTELLIGENT QUERY ENGINE
# Trainer or client asks any natural language question
# Claude reads Excel data and answers conversationally
# ══════════════════════════════════════════════════════

def read_excel_as_text(filepath):
    """Read all rows from Excel and return as structured text for Claude"""
    try:
        wb   = openpyxl.load_workbook(filepath)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return "No data found."

        headers = [str(h) for h in rows[0]]
        lines   = ["Columns: " + ", ".join(headers)]

        for row in rows[1:]:
            if any(cell for cell in row):
                line = " | ".join([str(v) if v else "-" for v in row])
                lines.append(line)

        return "\n".join(lines)
    except Exception as e:
        return "Error reading file: " + str(e)

def read_all_clients_summary():
    """Read all client Excel files and build a summary for cross-client queries"""
    mapping = load_mapping()
    ids     = mapping.get("ids", {})

    if not ids:
        return "No clients registered yet."

    summary = "ALL CLIENTS SUMMARY\n" + "="*40 + "\n"

    for cid, info in sorted(ids.items()):
        fp = os.path.join(DESKTOP_PATH, info.get("filename",""))
        if not os.path.exists(fp):
            continue

        try:
            wb   = openpyxl.load_workbook(fp)
            ws   = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows = [r for r in rows if r[2]]  # filter empty rows

            if not rows:
                continue

            first  = rows[0]
            latest = rows[-1]

            summary += "\nClient: " + info["name"] + " (" + cid + ")\n"
            summary += "Registered: " + info.get("registered","N/A") + "\n"
            summary += "Total entries: " + str(len(rows)) + "\n"
            # Dynamic column lookup
            hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
            def col(name): return next((i for i,h in enumerate(hdr) if h and name.lower() in str(h).lower()), None)
            wi = col('weight'); di = col('date'); ai = col('age'); hi2 = col('height')
            bgi = col('blood'); dti = col('diet'); gi = col('goal')

            summary += "Start weight: " + (str(first[wi]) if wi is not None else "N/A") + " on " + (str(first[di]) if di is not None else "N/A") + "\n"
            summary += "Latest weight: " + (str(latest[wi]) if wi is not None else "N/A") + " on " + (str(latest[di]) if di is not None else "N/A") + "\n"
            summary += "Age: " + (str(latest[ai]) if ai is not None else "N/A") + "\n"
            summary += "Height: " + (str(latest[hi2]) if hi2 is not None else "N/A") + "\n"
            summary += "Blood group: " + (str(latest[bgi]) if bgi is not None else "N/A") + "\n"
            summary += "Diet: " + (str(latest[dti]) if dti is not None else "N/A") + "\n"
            summary += "Goal: " + (str(latest[gi]) if gi is not None else "N/A") + "\n"
            summary += "Phone: " + info.get("phone","N/A") + "\n"

        except Exception as e:
            summary += "\nClient: " + info["name"] + " — Error reading file\n"

    return summary

def answer_query_with_claude(question, data, client_name=""):
    """Send question + Excel data to Claude and get a natural language answer"""
    context = "Client: " + client_name + "\n\n" if client_name else ""
    prompt  = (
        "You are FitTrack, an intelligent fitness data assistant.\n\n"
        "Answer the following question using ONLY the data provided below.\n"
        "Be concise, friendly and specific. Use numbers from the data.\n"
        "If the data does not contain enough information to answer, say so clearly.\n\n"
        "QUESTION: " + question + "\n\n"
        + context +
        "DATA:\n" + data + "\n\n"
        "Answer in a friendly WhatsApp message style. Keep it under 200 words."
    )
    return ask_claude(prompt, max_tokens=400)

def detect_query_target(message):
    """Extract which client the trainer is asking about"""
    raw = ask_claude(
        "Extract the client identifier from this fitness query.\n\n"
        "Message: " + message + "\n\n"
        "The identifier could be a client ID (like RAH-001, CHE-001) or a client name (like Rahul, Chenna).\n"
        "If the question is about ALL clients or no specific client, return: all\n"
        "If a specific client is mentioned, return ONLY their ID or name.\n\n"
        "Reply with ONE word or phrase only. Examples:\n"
        "- RAH-001\n"
        "- Rahul\n"
        "- all\n"
        "No other text.",
        max_tokens=20
    )
    return raw.strip().lower()

# ── Trainer intelligent query handler ──
def handle_trainer_query(question, trainer):
    """Handle any natural language question from trainer about client data"""
    print("Trainer query: " + question)
    send_text(trainer, "Looking that up for you... please wait!")

    target = detect_query_target(question)
    print("Query target: " + target)

    if target == "all":
        # Cross-client question
        data   = read_all_clients_summary()
        answer = answer_query_with_claude(question, data)
    else:
        # Specific client question
        fp, fn, cid, name = find_client(target)
        if not fp:
            send_text(trainer,
                "Could not find client: " + target + "\n"
                "Use trainer list to see all clients."
            )
            return
        data   = read_excel_as_text(fp)
        answer = answer_query_with_claude(question, data, name + " (" + cid + ")")

    send_text(trainer, answer)
    print("Query answered for trainer")

# ── Client intelligent query handler ──
def handle_client_query(question, sender):
    """Handle any natural language question from registered client about their own data"""
    fp, fn, cid, name = get_client_by_phone(sender)
    if not fp:
        send_text(sender, "Profile not found. Please contact your trainer.")
        return

    print("Client query from " + name + ": " + question)
    send_text(sender, "Let me check that for you... please wait!")

    data   = read_excel_as_text(fp)
    answer = answer_query_with_claude(question, data, name + " (" + cid + ")")
    send_text(sender, answer)
    print("Query answered for client: " + name)



# ══════════════════════════════════════════════════════
# MODULE 7: OCR / VISION — IMAGE PROCESSING
# Handles: weight scale photos, blood reports, 
#          body measurement charts, progress screenshots
# Auto-detects image type using Claude Vision
# ══════════════════════════════════════════════════════

def download_image(url):
    """Download image from UltraMsg URL and return as base64"""
    try:
        response = req.get(url, timeout=15)
        if response.status_code == 200:
            b64 = base64.b64encode(response.content).decode('utf-8')
            # Detect image type from content
            content_type = response.headers.get('content-type', 'image/jpeg')
            if 'png' in content_type:
                media_type = 'image/png'
            elif 'pdf' in content_type:
                media_type = 'application/pdf'
            elif 'webp' in content_type:
                media_type = 'image/webp'
            else:
                media_type = 'image/jpeg'
            print("Image downloaded: " + str(len(response.content)) + " bytes, type: " + media_type)
            return b64, media_type
        else:
            print("Image download failed: " + str(response.status_code))
            return None, None
    except Exception as e:
        print("Image download error: " + str(e))
        return None, None

def analyse_image_with_claude(b64_image, media_type, caption=""):
    """Send image to Claude Vision and extract ALL fitness data"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    prompt = (
        "You are a fitness data extraction assistant with OCR capabilities.\n\n"
        "Analyse this image and extract ALL relevant fitness data.\n\n"
        "The image may be one of these types:\n"
        "1. WEIGHT SCALE — extract the weight reading shown on display\n"
        "2. BLOOD REPORT — extract all test values (HB, sugar, BP, cholesterol, etc)\n"
        "3. BODY MEASUREMENT CHART — extract chest, waist, hips, etc\n"
        "4. PROGRESS SCREENSHOT — extract weight/measurement values and dates\n"
        "5. FITNESS REPORT — extract any health/fitness metrics visible\n\n"
        + ("Caption provided by sender: " + caption + "\n\n" if caption else "") +
        "Rules:\n"
        "1. First identify the image type\n"
        "2. Extract ALL numeric values with their labels and units\n"
        "3. Use clean key names with underscores\n"
        "4. Include units in values (82kg, 120/80 mmHg, 11.5 g/dL etc)\n"
        "5. If a date is visible in the image extract it as _report_date\n"
        "6. Add _image_type field describing what the image is\n"
        "7. Add _summary field with a brief human readable summary\n\n"
        "Reply ONLY with JSON. No markdown. No backticks.\n"
        "Examples:\n\n"
        "Weight scale:\n"
        "{\"_image_type\":\"weight_scale\",\"weight\":\"79.5kg\","
        "\"_summary\":\"Weight scale showing 79.5 kg\"}\n\n"
        "Blood report:\n"
        "{\"_image_type\":\"blood_report\","
        "\"hemoglobin\":\"13.5 g/dL\",\"blood_sugar_fasting\":\"95 mg/dL\","
        "\"total_cholesterol\":\"180 mg/dL\",\"_report_date\":\"05-May-2026\","
        "\"_summary\":\"Blood report — all values within normal range\"}\n\n"
        "Body measurements:\n"
        "{\"_image_type\":\"body_measurements\","
        "\"chest\":\"42 inches\",\"waist\":\"34 inches\",\"hips\":\"38 inches\","
        "\"_summary\":\"Body measurement chart showing chest 42, waist 34, hips 38\"}"
    )

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64_image
                        }
                    },
                    {
                        "type": "text",
                        "text": prompt
                    }
                ]
            }]
        )
        raw = response.content[0].text.strip()
        raw = raw.replace('```json','').replace('```','').strip()
        print("Claude Vision response: " + raw)
        return json.loads(raw)
    except Exception as e:
        print("Claude Vision error: " + str(e))
        return None

def handle_image_message(msg_data, sender):
    """Handle incoming image/media from WhatsApp.
    Downloads image, sends to Claude Vision, updates Excel."""

    # Get image URL from UltraMsg message data
    media_url = msg_data.get('media', '')
    caption   = msg_data.get('body', '').strip()

    if not media_url:
        print("No media URL found in message")
        return

    print("Image received from: " + sender)
    print("Media URL: " + media_url)
    print("Caption: " + caption)

    # Find client record
    if is_trainer(sender):
        # Trainer sending image — need to figure out which client
        # Check if caption mentions a client
        if caption:
            parsed = ask_claude(
                "Extract client ID or name from this trainer message caption.\n\n"
                "Caption: " + caption + "\n\n"
                "If a client ID (like RAH-001) or name is mentioned, return it.\n"
                "If no client mentioned, return: none\n\n"
                "Reply with client ID, name, or 'none' only.",
                max_tokens=20
            ).strip().lower()

            if parsed != 'none' and parsed:
                fp, fn, cid, name = find_client(parsed)
            else:
                # No client mentioned — ask trainer to specify
                send_text(sender,
                    "Please specify which client this image is for.\n\n"
                    "Example: Send the image with caption:\n"
                    "RAH-001 blood report\n"
                    "or\n"
                    "Rahul weight update"
                )
                return
        else:
            send_text(sender,
                "Please add a caption with the client name or ID.\n\n"
                "Example:\n"
                "RAH-001 blood report\n"
                "Rahul weight scale"
            )
            return
    else:
        # Client sending their own image
        fp, fn, cid, name = get_client_by_phone(sender)

    if not fp:
        send_text(sender,
            "Profile not found.\n"
            + ("Please contact your trainer." if not is_trainer(sender) else "Client not found — check ID or name.")
        )
        return

    send_text(sender, "Analysing your image... please wait! 🔍")

    # Download image
    b64, media_type = download_image(media_url)
    if not b64:
        send_text(sender, "Could not download the image. Please try again.")
        return

    # Analyse with Claude Vision
    extracted = analyse_image_with_claude(b64, media_type, caption)
    if not extracted:
        send_text(sender, "Could not read the image. Please make sure the image is clear and try again.")
        return

    # Get metadata
    image_type  = extracted.pop('_image_type', 'image')
    summary     = extracted.pop('_summary', 'Image processed')
    report_date = extracted.pop('_report_date', '')
    note_text   = "Image: " + image_type.replace('_', ' ').title()
    if report_date:
        note_text += " dated " + report_date

    # Remove internal fields
    extracted.pop('_comment', '')

    # Determine row type
    if is_trainer(sender):
        row_type   = "Trainer Update"
        updated_by = "Trainer (Image)"
    else:
        row_type   = "Client Update"
        updated_by = "Client (Image)"

    # Update Excel with extracted data
    add_row(fp, extracted, row_type, note=note_text, updated_by=updated_by)

    # Build confirmation message
    cap_now = datetime.now().strftime("%d %b %Y")
    reply = (
        "Image processed successfully! 🎉\n\n"
        + "Type: " + image_type.replace('_', ' ').title() + "\n"
        + "Client: " + name + " (" + cid + ")\n\n"
        + "Extracted data:\n"
    )

    for key, val in extracted.items():
        reply += "• " + key.replace('_', ' ').title() + ": " + str(val) + "\n"

    if report_date:
        reply += "\nReport date: " + report_date + "\n"

    reply += "\n" + summary + "\n\nExcel updated! ✅"

    send_text(sender, reply)

    # Notify other party
    if is_trainer(sender):
        # Notify client
        client_phone = get_phone_for_client(cid)
        if client_phone:
            send_text(client_phone,
                "Your trainer uploaded your " + image_type.replace('_',' ') + " report!\n\n"
                + summary + "\n\n"
                + "Send 'my details' to see your updated profile."
            )
    else:
        # Notify trainer
        send_to_trainer(
            "Client image received!\n\n"
            + "Client: " + name + " (" + cid + ")\n"
            + "Type: " + image_type.replace('_',' ').title() + "\n"
            + summary + "\n\n"
            + "Excel updated automatically!"
        )

    print("Image processed and Excel updated for: " + name + " (" + cid + ")")


# ══════════════════════════════════════════════════════
# WEBHOOK — MAIN ENTRY POINT
# ══════════════════════════════════════════════════════

@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        data = request.get_json(force=True, silent=True) or request.form.to_dict()
        print("\n" + "="*50)
        print("FITTRACK WEBHOOK")
        print("="*50)

        msg      = data.get('data', data)
        body     = msg.get('body', '')
        sender   = msg.get('from', '')
        from_me  = msg.get('fromMe', False)
        msg_type = msg.get('type', 'chat')
        media    = msg.get('media', '')

        print("From: "    + sender)
        print("Body: "    + body)
        print("Type: "    + msg_type)
        print("Media: "   + str(bool(media)))

        # ── Guard clauses ──
        if not sender:
            return 'ok', 200
        if from_me is True or str(from_me).lower() == 'true':
            return 'ok', 200
        # Only block the UltraMsg bot number (OWN_NUMBER)
        sender_clean = clean_number(sender)
        own_clean    = clean_number(OWN_NUMBER)
        if sender_clean == own_clean or own_clean in sender_clean or sender_clean in own_clean:
            print("Skipping own bot number: " + sender_clean)
            return 'ok', 200

        # ── IMAGE / MEDIA MESSAGE ──
        # Route to image handler if:
        # 1. Message type is image/document/video
        # 2. AND media URL is actually present (Webhook Download Media must be ON)
        if msg_type in ['image', 'document', 'video'] and media:
            print("Media message with URL — routing to image handler")
            handle_image_message(msg, sender)
            return 'ok', 200

        # Image type but no media URL — Webhook Download Media is OFF
        if msg_type in ['image', 'document', 'video'] and not media:
            print("Image received but no media URL — Webhook Download Media is OFF in UltraMsg")
            send_text(sender,
                "Image received but could not be downloaded.\n\n"
                "Please ask your admin to enable \'Webhook Download Media\' in UltraMsg settings.\n\n"
                "Alternatively send the data as a text message:\n"
                "update weight 79kg"
            )
            return 'ok', 200

        # ── Text message guard — must have body ──
        if not body:
            return 'ok', 200

        # ── TRAINER — identified by phone number, no prefix needed ──
        # If sender IS a trainer → handle all their messages
        if is_trainer(sender):
            if body.strip().lower().startswith('trainer'):
                # Structured trainer command — register, update, note, list, details
                handle_trainer(body, sender)
            else:
                # Free form message from trainer — treat as intelligent query
                handle_trainer_query(body, sender)
            return 'ok', 200

        # ── CLIENT — must be registered ──
        if not is_registered(sender):
            send_text(sender,
                "Welcome to FitTrack!\n\n"
                "You are not registered yet.\n\n"
                "Please contact your trainer to get registered.\n"
                "Your trainer will set up your profile and send you your unique FitTrack ID."
            )
            return 'ok', 200

        # ── REGISTERED CLIENT ──
        intent = detect_client_intent(body)

        if intent == "update":
            client_update(body, sender)
        elif intent == "my_details":
            client_details(sender)
        elif intent == "progress":
            client_progress(sender)
        elif intent == "query":
            handle_client_query(body, sender)
        else:
            send_text(sender,
                "FitTrack Commands:\n\n"
                "UPDATE your details:\n"
                "update weight 79kg\n"
                "update goal muscle gain\n"
                "update diet keto\n\n"
                "GET YOUR FILE:\n"
                "my details\n\n"
                "SEE YOUR PROGRESS:\n"
                "my progress\n\n"
                "ASK ANYTHING:\n"
                "how much weight have I lost?\n"
                "what was my starting weight?\n"
                "am I on track for my goal?"
            )

        print("="*50 + "\n")
        return 'ok', 200

    except Exception as e:
        print("ERROR: " + str(e))
        import traceback; traceback.print_exc()
        return 'error', 500

# ══════════════════════════════════════════════════════
# UTILITY ROUTES
# ══════════════════════════════════════════════════════

@app.route('/check_payments', methods=['GET'])
def check_all_payments():
    """Manual trigger to check payment dues for all clients"""
    m   = load_mapping()
    ids = m.get("ids", {})
    checked = 0
    for cid, info in ids.items():
        fp = os.path.join(DESKTOP_PATH, info.get("filename",""))
        cp = info.get("phone","")
        if fp and os.path.exists(fp) and cp:
            check_payment_due(fp, cid, info["name"], cp)
            checked += 1
    return "Payment check done for " + str(checked) + " clients!"

@app.route('/', methods=['GET'])
def home():
    return """
    <h2>FitTrack Bot v14 — Test Panel</h2>
    <p>Bot is running!</p>
    <hr>
    <h3>Test Query Engine</h3>
    <form method="GET" action="/test_query">
        <label>Client ID or Name:</label><br>
        <input type="text" name="id" placeholder="CHE-001 or Chenna or all" style="width:300px; padding:6px; margin:6px 0"><br>
        <label>Your Question:</label><br>
        <input type="text" name="q" placeholder="show last 5 weights" style="width:300px; padding:6px; margin:6px 0"><br><br>
        <input type="submit" value="Ask Claude" style="padding:8px 20px; background:#028090; color:white; border:none; cursor:pointer">
    </form>
    <hr>
    <h3>Quick Test Links</h3>
    <ul>
        <li><a href="/test_query?id=all&q=how many clients do I have">How many clients?</a></li>
        <li><a href="/test_query?id=all&q=which client has lost most weight">Who lost most weight?</a></li>
        <li><a href="/test_query?id=CHE-001&q=show last 5 weights">Last 5 weights for CHE-001</a></li>
        <li><a href="/test_query?id=CHE-001&q=how much weight has this client lost">Weight lost for CHE-001</a></li>
        <li><a href="/test_query?id=all&q=list all clients with their current weight">All clients current weight</a></li>
        <li><a href="/clients">View all registered clients</a></li>
    </ul>
    """

@app.route('/test_query', methods=['GET'])
def test_query():
    """Test the query engine without using WhatsApp messages"""
    question  = request.args.get('q', '').strip()
    client_id = request.args.get('id', '').strip()

    if not question:
        return "<h3>Error</h3><p>Please provide a question using ?q=your question</p><p><a href='/'>Back</a></p>"

    # Build response
    result = "<h2>FitTrack Query Test</h2>"
    result += "<p><b>Question:</b> " + question + "</p>"
    result += "<p><b>Client:</b> " + (client_id or "not specified") + "</p>"
    result += "<hr>"

    try:
        if not client_id or client_id.lower() == "all":
            data   = read_all_clients_summary()
            answer = answer_query_with_claude(question, data)
        else:
            fp, fn, cid, name = find_client(client_id)
            if not fp:
                return "<h3>Client not found: " + client_id + "</h3><p><a href='/'>Back</a></p>"
            data   = read_excel_as_text(fp)
            answer = answer_query_with_claude(question, data, name + " (" + cid + ")")

        result += "<h3>Claude AI Answer:</h3>"
        result += "<div style='background:#f0fafa; padding:15px; border-left:4px solid #028090; white-space:pre-wrap; font-family:Arial; line-height:1.6'>" + answer + "</div>"
        result += "<hr><p><a href='/'>Ask another question</a></p>"

    except Exception as e:
        result += "<h3>Error:</h3><p>" + str(e) + "</p>"
        result += "<p><a href='/'>Back</a></p>"

    return result

@app.route('/clients', methods=['GET'])
def list_clients():
    m   = load_mapping()
    ids = m.get("ids", {})
    out = "<h2>FitTrack Clients (" + str(len(ids)) + " total)</h2><ul>"
    for cid, info in sorted(ids.items()):
        out += "<li><b>" + cid + "</b> — " + info["name"] + " — Joined: " + info.get("registered","") + "</li>"
    out += "</ul><p><a href='/'>Back to Test Panel</a></p>"
    return out

# ══════════════════════════════════════════════════════
# START
# ══════════════════════════════════════════════════════

if __name__ == '__main__':
    print("="*50)
    print("FitTrack Bot v14 — Full Featured + OCR + Payments")
    print("Registration : Trainer ONLY")
    print("Updates      : Trainer + Client")
    print("ID Format    : RAH-001, PRI-001, RAH-002")
    print("Folder       : " + DESKTOP_PATH)
    print("Webhook      : http://localhost:5000/webhook")
    print("Clients      : http://localhost:5000/clients")
    print("="*50)
    app.run(host='0.0.0.0', port=5000, debug=True)